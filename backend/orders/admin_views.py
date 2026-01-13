from io import BytesIO
from decimal import Decimal
import datetime

from django.db.models import Q, Sum, Count
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import status as drf_status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ViewSet

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import Order, OrderStatusHistory
from .admin_serializers import (
    AdminOrderListSerializer,
    AdminOrderDetailSerializer,
)


class AdminOrderPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = "page_size"
    max_page_size = 100


def _money(val):
    try:
        return f"{Decimal(val):,.2f}"
    except Exception:
        try:
            return f"{Decimal(str(val)):,.2f}"
        except Exception:
            return "0.00"


def _parse_date(s: str):
    """
    Parse YYYY-MM-DD safely.
    Returns date or None.
    """
    if not s:
        return None
    s = str(s).strip()
    try:
        return datetime.date.fromisoformat(s)
    except Exception:
        return None


def _range_from_query(params):
    """
    start/end are YYYY-MM-DD.
    Returns (start_dt, end_dt_exclusive) in timezone-aware datetimes or (None, None) if invalid/missing.
    """
    start_d = _parse_date(params.get("start"))
    end_d = _parse_date(params.get("end"))

    if not start_d and not end_d:
        return None, None

    # If one side missing, make it a single-day range
    if start_d and not end_d:
        end_d = start_d
    if end_d and not start_d:
        start_d = end_d

    # Normalize to [start, end+1day)
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(datetime.datetime.combine(start_d, datetime.time.min), tz)
    end_dt = timezone.make_aware(datetime.datetime.combine(end_d + datetime.timedelta(days=1), datetime.time.min), tz)
    return start_dt, end_dt


def build_invoice_pdf(order: Order) -> bytes:
    """
    Generates a clean PDF invoice for an Order and returns raw PDF bytes.
    """
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=36,
        leftMargin=36,
        topMargin=36,
        bottomMargin=36,
        title=f"Invoice {order.order_number}",
        author="Urban Cart",
    )

    styles = getSampleStyleSheet()
    story = []

    # Header
    story.append(Paragraph("<b>Urban Cart</b>", styles["Title"]))
    story.append(Paragraph("Invoice", styles["Heading2"]))
    story.append(Spacer(1, 10))

    # Meta block
    created = order.created_at.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d %H:%M")
    meta_lines = [
        f"<b>Invoice for:</b> {order.order_number}",
        f"<b>Order ID:</b> {order.id}",
        f"<b>Order Date:</b> {created}",
        f"<b>Status:</b> {order.status}",
        f"<b>Payment:</b> {order.payment_method} ({order.payment_status})",
    ]
    for line in meta_lines:
        story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1, 12))

    # Shipping / Customer info
    story.append(Paragraph("<b>Shipping Details</b>", styles["Heading3"]))
    ship_lines = [
        f"<b>Name:</b> {order.shipping_name}",
        f"<b>Phone:</b> {order.phone}",
        f"<b>City:</b> {order.city}",
        f"<b>Address:</b> {order.address}",
    ]
    if order.note:
        ship_lines.append(f"<b>Note:</b> {order.note}")
    for line in ship_lines:
        story.append(Paragraph(line, styles["BodyText"]))
    story.append(Spacer(1, 14))

    # Items table
    story.append(Paragraph("<b>Items</b>", styles["Heading3"]))

    data = [["#", "Product", "SKU", "Qty", "Unit Price", "Line Total"]]
    items = list(order.items.all())
    for idx, it in enumerate(items, start=1):
        data.append([
            str(idx),
            (it.name or ""),
            (it.sku or ""),
            str(it.quantity),
            _money(it.price),
            _money(it.line_total),
        ])

    table = Table(data, colWidths=[24, 220, 90, 40, 80, 80])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),

        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    story.append(Spacer(1, 14))

    # Totals table
    totals = [
        ["Subtotal", _money(order.subtotal)],
        ["Discount", f"-{_money(order.discount_total)}"],
        ["Shipping", _money(order.shipping_fee)],
        ["Total", _money(order.total)],
    ]
    totals_table = Table(totals, colWidths=[120, 120])
    totals_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -2), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -2), 10),

        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 11),

        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("LINEABOVE", (0, -1), (-1, -1), 0.75, colors.black),
        ("TOPPADDING", (0, -1), (-1, -1), 8),
    ]))
    story.append(totals_table)
    story.append(Spacer(1, 16))

    # Footer
    story.append(Paragraph("Thank you for shopping with Urban Cart.", styles["Italic"]))

    doc.build(story)
    pdf = buf.getvalue()
    buf.close()
    return pdf


class AdminOrderViewSet(ViewSet):
    permission_classes = [IsAuthenticated]

    # ------------------------
    # Helpers
    # ------------------------
    def _ensure_admin(self, request):
        user = request.user
        if not user or not user.is_authenticated:
            return Response({"detail": "Authentication required."}, status=401)

        if not user.is_staff and not user.is_superuser:
            return Response({"detail": "Admin access required."}, status=403)

        return None

    def _base_queryset(self):
        return (
            Order.objects.select_related("user")
            .prefetch_related("items__product", "status_history")
            .order_by("-id")
        )

    # ------------------------
    # List Orders (ADMIN)
    # ------------------------
    def list(self, request):
        denied = self._ensure_admin(request)
        if denied:
            return denied

        qs = self._base_queryset()

        # Status filter
        status_value = request.query_params.get("status")
        if status_value and status_value.lower() != "all":
            qs = qs.filter(status=status_value)

        # Payment filter
        payment_status = request.query_params.get("payment_status")
        if payment_status and payment_status.lower() != "all":
            qs = qs.filter(payment_status=payment_status)

        # ✅ SEARCH (SAFE)
        search = (request.query_params.get("search") or "").strip()
        if search:
            q = (
                Q(order_number__icontains=search)     # UC-000006
                | Q(shipping_name__icontains=search)  # customer name
                | Q(phone__icontains=search)          # phone
                | Q(user__email__icontains=search)    # email
            )

            # If numeric → also match DB ID
            if search.isdigit():
                q |= Q(id=int(search))

            qs = qs.filter(q)

        paginator = AdminOrderPagination()
        page = paginator.paginate_queryset(qs, request)

        serializer = AdminOrderListSerializer(
            page, many=True, context={"request": request}
        )
        return paginator.get_paginated_response(serializer.data)

    # ------------------------
    # Retrieve Single Order
    # ------------------------
    def retrieve(self, request, pk=None):
        denied = self._ensure_admin(request)
        if denied:
            return denied

        try:
            order = self._base_queryset().get(pk=pk)
        except Order.DoesNotExist:
            return Response({"detail": "Order not found."}, status=404)

        serializer = AdminOrderDetailSerializer(
            order, context={"request": request}
        )
        return Response(serializer.data)

    # ------------------------
    # Update Order (status/payment)
    # ------------------------
    def partial_update(self, request, pk=None):
        denied = self._ensure_admin(request)
        if denied:
            return denied

        try:
            order = (
                Order.objects.select_related("user")
                .prefetch_related("items__product", "status_history")
                .get(pk=pk)
            )
        except Order.DoesNotExist:
            return Response({"detail": "Order not found."}, status=404)

        data = request.data or {}

        new_status = data.get("status")
        note = (data.get("note") or "").strip()
        payment_status = data.get("payment_status")
        payment_method = data.get("payment_method")

        old_status = order.status

        if payment_status is not None:
            order.payment_status = payment_status

        if payment_method is not None:
            order.payment_method = payment_method

        if new_status and str(new_status).strip():
            order.status = new_status

        order.updated_at = timezone.now()
        order.save()

        # Status history
        if new_status and new_status != old_status:
            OrderStatusHistory.objects.create(
                order=order,
                status=new_status,
                changed_by=request.user,
                note=note,
            )

        refreshed = (
            Order.objects.select_related("user")
            .prefetch_related("items__product", "status_history")
            .get(pk=order.pk)
        )

        serializer = AdminOrderDetailSerializer(
            refreshed, context={"request": request}
        )
        return Response(serializer.data, status=drf_status.HTTP_200_OK)

    # ------------------------
    # ✅ Download Invoice PDF (ADMIN)
    # ------------------------
    @action(detail=True, methods=["get"], url_path="invoice")
    def invoice(self, request, pk=None):
        denied = self._ensure_admin(request)
        if denied:
            return denied

        try:
            order = (
                Order.objects.select_related("user")
                .prefetch_related("items")
                .get(pk=pk)
            )
        except Order.DoesNotExist:
            return Response({"detail": "Order not found."}, status=404)

        pdf_bytes = build_invoice_pdf(order)
        filename = f"invoice-{order.order_number}.pdf"

        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp

    # ------------------------
    # ✅ NEW: Analytics (ADMIN)
    # GET /api/admin/orders/analytics/?start=YYYY-MM-DD&end=YYYY-MM-DD
    # ------------------------
    @action(detail=False, methods=["get"], url_path="analytics")
    def analytics(self, request):
        denied = self._ensure_admin(request)
        if denied:
            return denied

        qs = Order.objects.all()

        # Date range (optional)
        start_dt, end_dt = _range_from_query(request.query_params)
        if start_dt and end_dt:
            qs_range = qs.filter(created_at__gte=start_dt, created_at__lt=end_dt)
        else:
            qs_range = qs.none()

        # "Today" revenue/orders (based on server timezone)
        now = timezone.localtime()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        tomorrow_start = today_start + datetime.timedelta(days=1)

        qs_today = qs.filter(created_at__gte=today_start, created_at__lt=tomorrow_start)

        # "This month" revenue/orders
        month_start = today_start.replace(day=1)
        if month_start.month == 12:
            next_month_start = month_start.replace(year=month_start.year + 1, month=1)
        else:
            next_month_start = month_start.replace(month=month_start.month + 1)

        qs_month = qs.filter(created_at__gte=month_start, created_at__lt=next_month_start)

        # Revenue definitions:
        # - revenue_all: sum(total) for non-cancelled orders
        # - revenue_paid: sum(total) for payment_status=paid and non-cancelled
        non_cancelled = ~Q(status__in=[Order.Status.CANCELLED, Order.Status.REFUNDED])

        def _sum_total(qs_in):
            return qs_in.aggregate(v=Sum("total"))["v"] or Decimal("0.00")

        def _sum_total_paid(qs_in):
            return (
                qs_in.filter(payment_status=Order.PaymentStatus.PAID)
                .aggregate(v=Sum("total"))["v"]
                or Decimal("0.00")
            )

        # Today
        today_orders = qs_today.count()
        today_revenue_all = _sum_total(qs_today.filter(non_cancelled))
        today_revenue_paid = _sum_total_paid(qs_today.filter(non_cancelled))

        # Month
        month_orders = qs_month.count()
        month_revenue_all = _sum_total(qs_month.filter(non_cancelled))
        month_revenue_paid = _sum_total_paid(qs_month.filter(non_cancelled))

        # Range totals (if provided)
        range_orders = qs_range.count() if start_dt and end_dt else 0
        range_revenue_all = _sum_total(qs_range.filter(non_cancelled)) if start_dt and end_dt else Decimal("0.00")
        range_revenue_paid = _sum_total_paid(qs_range.filter(non_cancelled)) if start_dt and end_dt else Decimal("0.00")

        # Daily breakdown inside range (if provided)
        daily = []
        if start_dt and end_dt:
            daily_rows = (
                qs_range.filter(non_cancelled)
                .annotate(day=TruncDate("created_at"))
                .values("day")
                .annotate(
                    orders=Count("id"),
                    revenue_all=Sum("total"),
                    revenue_paid=Sum("total", filter=Q(payment_status=Order.PaymentStatus.PAID)),
                )
                .order_by("day")
            )
            for r in daily_rows:
                daily.append({
                    "date": r["day"].isoformat() if r["day"] else None,
                    "orders": int(r["orders"] or 0),
                    "revenue_all": str((r["revenue_all"] or Decimal("0.00"))),
                    "revenue_paid": str((r["revenue_paid"] or Decimal("0.00"))),
                })

        return Response({
            "timezone": str(timezone.get_current_timezone()),
            "today": {
                "orders": today_orders,
                "revenue_all": str(today_revenue_all),
                "revenue_paid": str(today_revenue_paid),
            },
            "this_month": {
                "orders": month_orders,
                "revenue_all": str(month_revenue_all),
                "revenue_paid": str(month_revenue_paid),
                "month_start": month_start.date().isoformat(),
            },
            "range": {
                "start": request.query_params.get("start"),
                "end": request.query_params.get("end"),
                "orders": range_orders,
                "revenue_all": str(range_revenue_all),
                "revenue_paid": str(range_revenue_paid),
                "daily_breakdown": daily,
            },
        }, status=200)

    # ------------------------
    # ✅ NEW: Export Excel (ADMIN)
    # GET /api/admin/orders/export-excel/?start=YYYY-MM-DD&end=YYYY-MM-DD
    # ------------------------
    @action(detail=False, methods=["get"], url_path="export-excel")
    def export_excel(self, request):
        denied = self._ensure_admin(request)
        if denied:
            return denied

        qs = (
            Order.objects.select_related("user")
            .order_by("-id")
        )

        start_dt, end_dt = _range_from_query(request.query_params)
        if start_dt and end_dt:
            qs = qs.filter(created_at__gte=start_dt, created_at__lt=end_dt)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders"

        headers = [
            "ID",
            "Order Number",
            "Created At",
            "Customer",
            "Email",
            "Phone",
            "City",
            "Status",
            "Payment Status",
            "Payment Method",
            "Subtotal",
            "Discount",
            "Shipping",
            "Total",
        ]
        ws.append(headers)

        tz = timezone.get_current_timezone()

        for o in qs.iterator():
            created = timezone.localtime(o.created_at, tz).strftime("%Y-%m-%d %H:%M:%S")
            ws.append([
                o.id,
                o.order_number,
                created,
                o.shipping_name,
                getattr(o.user, "email", "") if o.user_id else "",
                o.phone,
                o.city,
                o.status,
                o.payment_status,
                o.payment_method,
                float(o.subtotal or 0),
                float(o.discount_total or 0),
                float(o.shipping_fee or 0),
                float(o.total or 0),
            ])

        # Basic column sizing
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18

        # Write to bytes
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        filename = "orders-report.xlsx"
        resp = HttpResponse(
            out.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
